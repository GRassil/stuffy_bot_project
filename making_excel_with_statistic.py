from openpyxl import Workbook

from models import *


def make_excel():
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "ОГЭ"
    ws2 = wb.create_sheet("ПРОФИЛЬ")
    ws3 = wb.create_sheet("БАЗА")

    for ws in wb.sheetnames:
        wb[ws].append(
            [
                "Имя",
                "Фамилия",
                "Всего решено задач",
                "Всего правильно решённых задач",
                "Процент решённых задач",
                "Всего решено вариантов",
                "Средний бал",
                "Максимальный бал",
            ]
        )

    with conn:
        try:
            for user in User.select():
                examtype = ExamType.get(ExamType.exam_type == user.exam_type)
                data = {
                    "A": user.name,
                    "B": user.lastname,
                    "C": user.total_ex,
                    "D": user.correct_ex,
                    "E": round((int(user.correct_ex) / int(user.total_ex)) * 100, 2)
                    if int(user.total_ex) != 0
                    else 0,
                    "F": user.total_tests,
                    "G": round(
                        int(user.total_points)
                        / (int(user.total_tests) * int(examtype.max_points)),
                        2,
                    )
                    if int(user.total_tests) != 0
                    else 0,
                    "H": user.max_points_per_test,
                }
                if user.if_get_course != "0":
                    if user.exam_type_id == "ОГЭ":
                        ws1.append(data)
                    elif user.exam_type_id == "ПРОФИЛЬ":
                        ws2.append(data)
                    elif user.exam_type_id == "БАЗА":
                        ws3.append(data)
        except Exception as e:
            print(e)
    wb.save("statistic.xlsx")
